#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
投資ダッシュボード用のデータ収集スクリプト。

GitHub Actions から毎日実行され、data/toushi.json を更新する。

■ 設計方針
  1. 取得先はすべて無料・APIキー不要。
  2. 1つの系列に複数の取得先を用意し、上から順に試す（カスケード）。
     GitHub のサーバーからはボット判定で弾かれるサイトがあるため、
     公的機関（財務省・米財務省・BLS）を優先し、民間サイトは後ろに置く。
  3. どこかが壊れても全体は止めない。取れなかった系列は前回値を保持する。
  4. 何をどう試して何が起きたかを全部記録し、JSON の notes に残す。
     Actions のログを見なくても、data/toushi.json を見れば原因がわかる。
"""

import datetime as dt
import json
import os
import re
import sys
import time
from urllib.parse import quote

import requests

# ---------------------------------------------------------------------------
# 基本設定
# ---------------------------------------------------------------------------

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT_PATH = os.path.join(ROOT, "data", "toushi.json")

YEARS = 20
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": UA,
    "Accept": "text/csv,application/json,text/html;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.8,en;q=0.6",
})

TODAY = dt.date.today()
START = TODAY - dt.timedelta(days=365 * YEARS + 10)
CUTOFF = START.isoformat()

DIAG = []          # 取得の記録（成功も失敗も全部）
NOTES = []         # 画面に出す短いメモ

STATUS_PATH = os.path.join(ROOT, "data", "toushi-status.json")


def valid_date(s):
    """使ってよい日付か確かめる。
    ・ちゃんとした暦の日付か（2026-13-01 のようなものを弾く）
    ・20年より古くないか
    ・未来でないか（配布元が年末までの空行を持っていることがある）
    """
    try:
        d = dt.date.fromisoformat(s)
    except (ValueError, TypeError):
        return False
    return START <= d <= TODAY


def diag(line):
    print("    " + line, file=sys.stderr)
    DIAG.append(line)


def get(url, timeout=25, tries=2, **kw):
    """GET。失敗しても例外を投げずに None を返す。理由は呼び出し側で記録する。"""
    last = ""
    for i in range(tries):
        try:
            r = SESSION.get(url, timeout=timeout, **kw)
            if r.status_code == 200:
                return r
            last = "HTTP %d" % r.status_code
        except requests.exceptions.Timeout:
            last = "タイムアウト"
        except Exception as e:                       # noqa: BLE001
            last = type(e).__name__
        if i + 1 < tries:
            time.sleep(2)
    get.last_error = last
    return None


get.last_error = ""


def looks_like_botwall(text):
    """ボット判定ページが返ってきていないか見る。"""
    head = text[:600].lower()
    for sign in ("requires javascript", "enable javascript", "cf-browser-verification",
                 "just a moment", "captcha", "<!doctype html><html><head><meta charset"):
        if sign in head:
            return True
    return False


# ---------------------------------------------------------------------------
# 系列の定義
# ---------------------------------------------------------------------------

META = {
    "nikkei":      dict(label="日経平均株価",      unit="円",   group="price", expensive_high=None, source="Yahoo Finance ^N225"),
    "topix":       dict(label="TOPIX",             unit="pt",   group="price", expensive_high=None, source="Yahoo Finance ^TPX"),
    "spx":         dict(label="S&P500",            unit="pt",   group="price", expensive_high=None, source="Yahoo Finance ^GSPC"),
    "gold":        dict(label="金",                unit="$/oz", group="price", expensive_high=None, source="Yahoo Finance GC=F"),
    "silver":      dict(label="銀",                unit="$/oz", group="price", expensive_high=None, source="Yahoo Finance SI=F"),
    "usdjpy":      dict(label="ドル円",            unit="円",   group="price", expensive_high=None, source="Yahoo Finance JPY=X"),

    "nikkei_per":  dict(label="日経平均 PER",      unit="倍",   group="value", expensive_high=True,  source="日経平均プロフィル(加重平均)"),
    "nikkei_pbr":  dict(label="日経平均 PBR",      unit="倍",   group="value", expensive_high=True,  source="日経平均プロフィル(加重平均)"),
    "spx_per":     dict(label="S&P500 PER",        unit="倍",   group="value", expensive_high=True,  source="Shiller / multpl"),
    "cape":        dict(label="S&P500 CAPE",       unit="倍",   group="value", expensive_high=True,  source="Shiller CAPE"),

    "dgs10":       dict(label="米10年金利",        unit="%",    group="rate",  expensive_high=None, source="米財務省 イールドカーブ"),
    "real10":      dict(label="米10年 実質金利",   unit="%",    group="rate",  expensive_high=None, source="米財務省 実質イールドカーブ"),
    "bei10":       dict(label="10年 BEI",          unit="%",    group="rate",  expensive_high=None, source="名目10年 − 実質10年"),
    "jp10y":       dict(label="日本10年金利",      unit="%",    group="rate",  expensive_high=None, source="財務省 国債金利情報"),
    "us_cpi_yoy":  dict(label="米CPI 前年比",      unit="%",    group="rate",  expensive_high=None, source="BLS 公開API"),
    "jp_cpi_yoy":  dict(label="日本CPI 前年比",    unit="%",    group="rate",  expensive_high=None, source="FRED JPNCPIALLMINMEI"),

    "nt":          dict(label="NT倍率",            unit="倍",   group="ratio", expensive_high=None, source="日経平均 ÷ TOPIX"),
    "gsr":         dict(label="金銀比価",          unit="倍",   group="ratio", expensive_high=None, source="金 ÷ 銀"),
    "nikkei_usd":  dict(label="ドル建て日経",      unit="$",    group="ratio", expensive_high=None, source="日経平均 ÷ ドル円"),
    "nikkei_gold": dict(label="日経 ÷ 金",         unit="oz",   group="ratio", expensive_high=True,  source="ドル建て日経 ÷ 金価格"),
    "ys_jp":       dict(label="日本 イールドスプレッド", unit="%", group="spread", expensive_high=False, source="100÷日経PER − 日本10年金利"),
    "ys_us":       dict(label="米国 イールドスプレッド", unit="%", group="spread", expensive_high=False, source="100÷S&P500PER − 米10年金利"),
}

# 派生指標（計算で作るもの）。取得はしない。
DERIVED = ("bei10", "nt", "gsr", "nikkei_usd", "nikkei_gold", "ys_jp", "ys_us")


# ---------------------------------------------------------------------------
# 取得先 1: Yahoo Finance（株価指数・商品・為替）
#   チャートAPIを直接叩く。ライブラリ不要。
#   20年分は週次、直近6ヶ月は日次で取って重ねる。
# ---------------------------------------------------------------------------

def yahoo_chart(symbol, rng, interval):
    for host in ("query2", "query1"):
        url = ("https://%s.finance.yahoo.com/v8/finance/chart/%s?range=%s&interval=%s"
               % (host, quote(symbol), rng, interval))
        r = get(url, timeout=25, tries=1)
        if r is None:
            continue
        try:
            j = r.json()
        except Exception:                            # noqa: BLE001
            continue
        res = (j.get("chart") or {}).get("result") or []
        if not res:
            err = ((j.get("chart") or {}).get("error") or {}).get("description", "")
            if err:
                diag("      %s %s/%s: %s" % (symbol, rng, interval, err[:60]))
            continue
        res = res[0]
        stamps = res.get("timestamp") or []
        quotes = ((res.get("indicators") or {}).get("quote") or [{}])[0]
        closes = quotes.get("close") or []
        out = {}
        for i in range(min(len(stamps), len(closes))):
            if closes[i] is None:
                continue
            d = dt.datetime.utcfromtimestamp(stamps[i]).date().isoformat()
            out[d] = float(closes[i])
        if out:
            return out
    return {}


def fetch_yahoo(symbols):
    """候補シンボルを順に試し、週次(20年)＋日次(6ヶ月)を重ねて返す。"""
    for sym in symbols:
        weekly = yahoo_chart(sym, "20y", "1wk")
        if not weekly:
            weekly = yahoo_chart(sym, "max", "1wk")
        if not weekly:
            diag("      %s: 週次が取れず" % sym)
            continue
        daily = yahoo_chart(sym, "6mo", "1d")
        merged = dict(weekly)
        merged.update(daily)
        merged = {d: v for d, v in merged.items() if d >= CUTOFF}
        diag("      %s: 週次%d点 + 日次%d点 → %d点" % (sym, len(weekly), len(daily), len(merged)))
        return merged
    return {}


# ---------------------------------------------------------------------------
# 取得先 2: 米財務省（10年金利・10年実質金利）
#   年ごとのCSV。機械での取得を前提に公開されている。
# ---------------------------------------------------------------------------

def treasury_curve(kind, col_names):
    base = ("https://home.treasury.gov/resource-center/data-chart-center/"
            "interest-rates/daily-treasury-rates.csv/%d/all"
            "?type=%s&field_tdr_date_value=%d&page&_format=csv")
    out, ok_years, ng_years = {}, 0, []
    miss_streak = 0
    # 新しい年から順に取る。3年続けて取れなければ、その取得先は死んでいると判断して打ち切る
    for year in range(TODAY.year, START.year - 1, -1):
        if miss_streak >= 3:
            diag("      %s: 3年続けて取れないので打ち切り" % kind)
            break
        r = get(base % (year, kind, year), timeout=20, tries=1)
        if r is None:
            ng_years.append(str(year)); miss_streak += 1
            continue
        text = r.text.strip()
        if not text or looks_like_botwall(text):
            ng_years.append(str(year)); miss_streak += 1
            continue
        miss_streak = 0
        lines = text.splitlines()
        header = [h.strip().strip('"') for h in lines[0].split(",")]
        idx = None
        for want in col_names:
            if want in header:
                idx = header.index(want)
                break
        if idx is None:
            ng_years.append(str(year) + "(列不明)")
            continue
        for line in lines[1:]:
            cells = [c.strip().strip('"') for c in line.split(",")]
            if len(cells) <= idx or not cells[0]:
                continue
            try:
                m, d, y = cells[0].split("/")
                date = "%04d-%02d-%02d" % (int(y), int(m), int(d))
                out[date] = float(cells[idx])
            except (ValueError, IndexError):
                continue
        ok_years += 1
        time.sleep(0.4)
    if ng_years:
        diag("      取れなかった年: %s" % ",".join(ng_years[:8]))
    diag("      %s: %d年分OK / %d点" % (kind, ok_years, len(out)))
    return out


# ---------------------------------------------------------------------------
# 取得先 3: 財務省（日本の10年国債金利）
#   全期間分が1本のCSVにまとまっている。日付は和暦。
# ---------------------------------------------------------------------------

def fetch_jgb10y():
    url = "https://www.mof.go.jp/jgbs/reference/interest_rate/data/jgbcm_all.csv"
    r = get(url, timeout=45, tries=2)
    if r is None:
        return {}
    text = r.content.decode("cp932", errors="replace")
    out = {}
    era = {"S": 1925, "H": 1988, "R": 2018}          # 昭和 / 平成 / 令和
    for line in text.splitlines():
        cells = [c.strip() for c in line.split(",")]
        if len(cells) < 11:
            continue
        m = re.match(r"^([SHR])(\d+)\.(\d+)\.(\d+)$", cells[0])
        if not m:
            continue
        try:
            year = era[m.group(1)] + int(m.group(2))
            date = "%04d-%02d-%02d" % (year, int(m.group(3)), int(m.group(4)))
            if date < CUTOFF:
                continue
            out[date] = float(cells[10])             # 0=日付, 1=1年 … 10=10年
        except (ValueError, KeyError):
            continue
    return out


# ---------------------------------------------------------------------------
# 取得先 4: 日経平均プロフィル（日経の PER / PBR）
#   表の列は「日付 / 加重平均(倍) / 指数ベース(倍)」。加重平均を使う。
#   ※ここは前回の実行で正常に動いた実績がある。
# ---------------------------------------------------------------------------

def fetch_nikkei_ratio(kind):
    url = "https://indexes.nikkei.co.jp/nkave/archives/data?list=%s" % kind
    r = get(url, timeout=25, tries=2)
    if r is None:
        return {}
    r.encoding = r.apparent_encoding or "utf-8"
    html = r.text
    rows = re.findall(
        r"(\d{4})\.(\d{2})\.(\d{2})\s*</td>\s*<td[^>]*>\s*([\d.]+)\s*</td>\s*<td[^>]*>\s*([\d.]+)",
        html)
    if not rows:
        rows = re.findall(
            r"(\d{4})\.(\d{2})\.(\d{2})[^\d]{1,80}?([\d]+\.[\d]+)[^\d]{1,80}?([\d]+\.[\d]+)", html)
    out = {}
    for y, mo, d, weighted, _idx in rows:
        try:
            out["%s-%s-%s" % (y, mo, d)] = float(weighted)
        except ValueError:
            continue
    return out


# ---------------------------------------------------------------------------
# 取得先 5: Shiller のデータ（S&P500 の PER と CAPE）
#   イェール大 Shiller 教授が公開している月次データ。学術用なので機械取得に寛容。
#   ファイルは Excel。openpyxl で読む。
# ---------------------------------------------------------------------------

def _shiller_download():
    """shillerdata.com のページから ie_data.xls の場所を見つけて落とす。
    配布URLには版番号が付いていて更新のたびに変わるので、毎回ページから拾う。"""
    page = get("https://shillerdata.com/", timeout=30, tries=2)
    urls = []
    if page is not None:
        for href in re.findall(r'href="([^"]*ie_data\.xls[^"]*)"', page.text):
            urls.append(href.replace("&amp;", "&"))
    # ページから拾えなかったときの控え（2026-08 時点で有効だったURL）
    urls.append("https://img1.wsimg.com/blobby/go/e5e77e0b-59d1-44d9-ab25-4763ac982e53"
                "/downloads/e27e58c1-8ae0-488c-a976-a298708c7175/ie_data.xls")
    for u in urls:
        r = get(u, timeout=90, tries=1)
        if r is None:
            continue
        head = r.content[:4]
        # 中身が本当に Excel か、マジックナンバーで確かめる
        # （HTMLのエラーページやボット判定ページを掴まないように）
        if head[:2] == b"PK" or head == b"\xd0\xcf\x11\xe0":
            diag("      Shiller: %d KB 取得 (%s)" % (len(r.content) // 1024, u.split("/")[2]))
            return r.content
        diag("      Shiller: Excelではない応答 (%s)" % u.split("/")[2])
    return None


def _excel_rows(blob):
    """Excel を読んで、Data シートの行を list で返す。
    Shiller のファイルは旧形式(.xls)なので、新形式(.xlsx)と両方に対応する。"""
    import io
    if blob[:2] == b"PK":                            # .xlsx / .xlsm
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
        names = wb.sheetnames
        pick = next((n for n in names if n.strip().lower().startswith("data")), names[0])
        return list(wb[pick].iter_rows(values_only=True))

    import xlrd                                      # .xls（Excel 97-2003）
    book = xlrd.open_workbook(file_contents=blob)
    names = book.sheet_names()
    pick = next((n for n in names if n.strip().lower().startswith("data")), names[0])
    sheet = book.sheet_by_name(pick)
    return [tuple(sheet.row_values(i)) for i in range(sheet.nrows)]


def fetch_shiller():
    """Shiller の月次データから CAPE と 実績PER(P÷E) を取り出す。
    列の位置は決め打ちにせず、見出しの言葉で探し、値の妥当性でも確かめる。"""
    blob = _shiller_download()
    if blob is None:
        return {}

    try:
        rows = _excel_rows(blob)
    except Exception as e:                           # noqa: BLE001
        diag("      Shiller: Excelを開けず (%s: %s)" % (type(e).__name__, str(e)[:60]))
        return {}

    # --- 見出し行を探して列の位置を決める ---
    col_date = col_p = col_e = col_cape = None
    for row in rows[:14]:
        cells = [(" ".join(str(c).split())).lower() if c is not None else "" for c in row]
        if not any(c == "date" for c in cells):
            continue
        col_date = cells.index("date")
        for i, c in enumerate(cells):
            if not c:
                continue
            # CAPE の列。"Excess CAPE Yield" や "TR CAPE" は別物なので避ける
            if col_cape is None and "cape" in c and "excess" not in c and "yield" not in c and " tr " not in c:
                col_cape = i
            if col_p is None and "comp" in c:                      # "S&P Comp. P"
                col_p = i
            if col_e is None and "earnings" in c and "real" not in c and "scaled" not in c:
                col_e = i
        diag("      Shiller: 列 date=%s P=%s E=%s CAPE=%s"
             % (col_date, col_p, col_e, col_cape))
        break
    if col_date is None:
        diag("      Shiller: 見出し行が見つからず")
        return {}

    def cell(row, i):
        if i is None or i >= len(row):
            return None
        v = row[i]
        return float(v) if isinstance(v, (int, float)) else None

    per, cape = {}, {}
    for row in rows:
        d = cell(row, col_date)
        if d is None:
            continue
        # Shiller の日付は 1871.01 のような小数。小数第2位が月。
        year = int(d)
        month = int(round((d - year) * 100))
        if year < 1871 or not 1 <= month <= 12:
            continue
        date = "%04d-%02d-01" % (year, month)
        if date < CUTOFF:
            continue
        c = cell(row, col_cape)
        if c and 3 < c < 80:                      # CAPE がありえる範囲か確かめる
            cape[date] = round(c, 2)
        p, e = cell(row, col_p), cell(row, col_e)
        if p and e and e > 0:
            v = p / e
            if 3 < v < 120:                       # 実績PER がありえる範囲か確かめる
                per[date] = round(v, 2)

    diag("      Shiller: CAPE %d点 / 実績PER %d点" % (len(cape), len(per)))
    return {"cape": cape, "spx_per": per}


def fetch_multpl(path):
    """multpl.com の月次テーブル（Shiller が取れないときの控え）。"""
    url = "https://www.multpl.com/%s/table/by-month" % path
    r = get(url, timeout=25, tries=1)
    if r is None:
        return {}
    if looks_like_botwall(r.text):
        diag("      multpl: ボット判定ページが返ってきた")
        return {}
    rows = re.findall(
        r"<td[^>]*>\s*([A-Z][a-z]{2})\s+(\d{1,2}),\s*(\d{4})\s*</td>\s*<td[^>]*>\s*([\d.]+)",
        r.text)
    months = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
              "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
    out = {}
    for mon, day, year, val in rows:
        if mon in months:
            try:
                out["%s-%02d-%02d" % (year, months[mon], int(day))] = float(val)
            except ValueError:
                continue
    return out


# ---------------------------------------------------------------------------
# 取得先 6: BLS（米CPI）と FRED（控え／日本CPI）
# ---------------------------------------------------------------------------

def fetch_bls_cpi():
    """米CPI（都市部・全品目）の月次指数。10年ずつ2回に分けて取る。"""
    out = {}
    for y0 in (START.year, START.year + 10):
        y1 = min(y0 + 9, TODAY.year)
        if y0 > TODAY.year:
            break
        url = ("https://api.bls.gov/publicAPI/v1/timeseries/data/CUUR0000SA0"
               "?startyear=%d&endyear=%d" % (y0, y1))
        r = get(url, timeout=30, tries=2)
        if r is None:
            continue
        try:
            j = r.json()
        except Exception:                            # noqa: BLE001
            continue
        if j.get("status") != "REQUEST_SUCCEEDED":
            diag("      BLS: %s" % str(j.get("message"))[:80])
            continue
        for s in (j.get("Results") or {}).get("series", []):
            for row in s.get("data", []):
                try:
                    year = int(row["year"])
                    month = int(row["period"].replace("M", ""))
                    if not 1 <= month <= 12:
                        continue          # M13（年平均）は月次ではないので使わない
                    out["%04d-%02d-01" % (year, month)] = float(row["value"])
                except (ValueError, KeyError):
                    continue
        time.sleep(1)
    return out


def fetch_fred(series_id):
    url = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=%s" % series_id
    r = get(url, timeout=25, tries=1)
    if r is None:
        return {}
    text = r.text
    if looks_like_botwall(text):
        return {}
    out = {}
    for line in text.strip().splitlines()[1:]:
        cells = line.split(",")
        if len(cells) < 2:
            continue
        val = cells[1].strip()
        if val in (".", "", "NA"):
            continue
        try:
            out[cells[0].strip()] = float(val)
        except ValueError:
            continue
    return out


def to_yoy(monthly):
    """月次の指数から前年同月比(%)を作る。"""
    out = {}
    for date, val in monthly.items():
        y, m, d = date.split("-")
        prev = "%04d-%s-%s" % (int(y) - 1, m, d)
        if monthly.get(prev):
            out[date] = round((val / monthly[prev] - 1.0) * 100.0, 2)
    return out


# ---------------------------------------------------------------------------
# カスケード実行
# ---------------------------------------------------------------------------

def try_sources(key, candidates):
    """候補を上から試し、最初に中身が返ったものを採用する。"""
    label = META[key]["label"]
    for rank, (name, fn) in enumerate(candidates):
        t0 = time.time()
        try:
            data = fn()
        except Exception as e:                       # noqa: BLE001
            diag("  %s ← %s : 例外 %s (%s)" % (label, name, type(e).__name__, str(e)[:60]))
            continue
        sec = time.time() - t0
        if data:
            diag("  %s ← %s : OK %d点 (%s〜%s) %.1f秒"
                 % (label, name, len(data), min(data), max(data), sec))
            if rank > 0:                             # 控えの取得先を使ったときだけ書き換える
                META[key]["source"] = name + "（控え）"
                NOTES.append("%s は %s から取得しました" % (label, name))
            return data
        diag("  %s ← %s : 空 [%s] %.1f秒" % (label, name, get.last_error or "内容なし", sec))
    NOTES.append("%s を取得できませんでした（前回値を保持）" % label)
    return {}


def collect():
    raw = {}

    print("■ 価格（Yahoo Finance）")
    for key, syms in [
        ("nikkei", ["^N225"]),
        ("topix",  ["^TPX", "998405.T", "^TOPX"]),
        ("spx",    ["^GSPC"]),
        ("gold",   ["GC=F", "XAUUSD=X"]),
        ("silver", ["SI=F", "XAGUSD=X"]),
        ("usdjpy", ["JPY=X", "USDJPY=X"]),
    ]:
        raw[key] = try_sources(key, [
            ("Yahoo Finance", (lambda s=syms: fetch_yahoo(s))),
            ("FRED",          (lambda k=key: fetch_fred(
                {"nikkei": "NIKKEI225", "spx": "SP500", "usdjpy": "DEXJPUS"}.get(k, "")))),
        ])
        time.sleep(1)

    print("■ 米国の金利（米財務省）")
    raw["dgs10"] = try_sources("dgs10", [
        ("米財務省", lambda: treasury_curve("daily_treasury_yield_curve", ["10 Yr", "10 YR"])),
        ("FRED",     lambda: fetch_fred("DGS10")),
    ])
    raw["real10"] = try_sources("real10", [
        ("米財務省", lambda: treasury_curve("daily_treasury_real_yield_curve", ["10 YR", "10 Yr"])),
        ("FRED",     lambda: fetch_fred("DFII10")),
    ])

    print("■ 日本の金利（財務省）")
    raw["jp10y"] = try_sources("jp10y", [
        ("財務省 国債金利情報", fetch_jgb10y),
        ("FRED",               lambda: fetch_fred("IRLTLT01JPM156N")),
    ])

    print("■ 日経のバリュエーション（日経平均プロフィル）")
    raw["nikkei_per"] = try_sources("nikkei_per", [("日経平均プロフィル", lambda: fetch_nikkei_ratio("per"))])
    raw["nikkei_pbr"] = try_sources("nikkei_pbr", [("日経平均プロフィル", lambda: fetch_nikkei_ratio("pbr"))])

    print("■ S&P500 のバリュエーション")
    shiller = {}
    try:
        shiller = fetch_shiller()
    except Exception as e:                           # noqa: BLE001
        diag("  Shiller: 例外 %s" % type(e).__name__)
    raw["cape"] = try_sources("cape", [
        ("Shiller (Yale)", lambda: shiller.get("cape", {})),
        ("multpl",         lambda: fetch_multpl("shiller-pe")),
    ])
    raw["spx_per"] = try_sources("spx_per", [
        ("Shiller (Yale)", lambda: shiller.get("spx_per", {})),
        ("multpl",         lambda: fetch_multpl("s-p-500-pe-ratio")),
    ])

    print("■ インフレ")
    us_cpi = try_sources("us_cpi_yoy", [
        ("BLS 公開API", fetch_bls_cpi),
        ("FRED",        lambda: fetch_fred("CPIAUCSL")),
    ])
    raw["us_cpi_yoy"] = to_yoy(us_cpi)
    jp_cpi = try_sources("jp_cpi_yoy", [
        ("FRED", lambda: fetch_fred("JPNCPIALLMINMEI")),
    ])
    raw["jp_cpi_yoy"] = to_yoy(jp_cpi)

    return raw


# ---------------------------------------------------------------------------
# 統合・間引き・派生指標
# ---------------------------------------------------------------------------

def load_previous():
    if not os.path.exists(OUT_PATH):
        return {}
    try:
        with open(OUT_PATH, encoding="utf-8") as f:
            prev = json.load(f)
    except Exception as e:                           # noqa: BLE001
        diag("前回の JSON を読めなかった: %s" % e)
        return {}
    dates = prev.get("dates", [])
    out = {}
    for key, values in prev.get("series", {}).items():
        out[key] = {dates[i]: v for i, v in enumerate(values)
                    if v is not None and i < len(dates)}
    return out


def merge(previous, fresh):
    """前回値に今回ぶんを重ね、おかしな日付をここで一括して落とす。
    取得先ごとに気をつけるより、最後に1か所で弾くほうが漏れがない。"""
    merged, dropped = {}, 0
    for key in set(list(previous.keys()) + list(fresh.keys())):
        base = dict(previous.get(key, {}))
        base.update(fresh.get(key, {}))
        kept = {d: v for d, v in base.items() if valid_date(d)}
        dropped += len(base) - len(kept)
        merged[key] = kept
    if dropped:
        diag("おかしな日付を %d 点ぶん落とした（未来日・存在しない日付など）" % dropped)
    return merged


def latest_on_or_before(series, target):
    best = None
    for d in series:
        if d <= target and (best is None or d > best):
            best = d
    return series[best] if best is not None else None


def build_date_axis(all_dates):
    """直近3ヶ月=日次 / 2年前まで=週次 / 20年前まで=月次 に間引く。"""
    dates = sorted(set(all_dates))
    d90 = (TODAY - dt.timedelta(days=90)).isoformat()
    d2y = (TODAY - dt.timedelta(days=730)).isoformat()
    keep, seen_week, seen_month = set(), set(), set()
    for d in dates:
        if d >= d90:
            keep.add(d)
        elif d >= d2y:
            y, w, _ = dt.date.fromisoformat(d).isocalendar()
            if (y, w) not in seen_week:
                seen_week.add((y, w)); keep.add(d)
        else:
            if d[:7] not in seen_month:
                seen_month.add(d[:7]); keep.add(d)
    return sorted(keep)


def add_derived(table, dates):
    """比率・BEI・イールドスプレッドを計算して足す。"""
    g = lambda k: table.get(k, {})                   # noqa: E731
    # 前回計算ぶんを土台にする（材料が一時的に取れなくても値が消えないように）
    out = {k: dict(table.get(k, {})) for k in DERIVED}
    for d in dates:
        nk  = latest_on_or_before(g("nikkei"), d)
        tp  = latest_on_or_before(g("topix"), d)
        gd  = latest_on_or_before(g("gold"), d)
        sv  = latest_on_or_before(g("silver"), d)
        fx  = latest_on_or_before(g("usdjpy"), d)
        np_ = latest_on_or_before(g("nikkei_per"), d)
        sp_ = latest_on_or_before(g("spx_per"), d)
        j10 = latest_on_or_before(g("jp10y"), d)
        u10 = latest_on_or_before(g("dgs10"), d)
        r10 = latest_on_or_before(g("real10"), d)

        if u10 is not None and r10 is not None:
            out["bei10"][d] = round(u10 - r10, 2)    # 期待インフレ率
        if nk and tp:
            out["nt"][d] = round(nk / tp, 3)
        if gd and sv:
            out["gsr"][d] = round(gd / sv, 2)
        nk_usd = nk / fx if (nk and fx) else None
        if nk_usd:
            out["nikkei_usd"][d] = round(nk_usd, 2)
            if gd:
                out["nikkei_gold"][d] = round(nk_usd / gd, 4)
        if np_ and j10 is not None:
            out["ys_jp"][d] = round(100.0 / np_ - j10, 2)
        if sp_ and u10 is not None:
            out["ys_us"][d] = round(100.0 / sp_ - u10, 2)
    table.update(out)
    return table


def main():
    t0 = time.time()
    fresh = collect()
    table = merge(load_previous(), fresh)

    # 日付軸は「点数が多い系列すべて」から作る。
    # 価格だけに頼ると、価格の取得先が落ちたとき軸が1点に潰れてしまう。
    axis_source = []
    for key, col in table.items():
        if key in DERIVED:
            continue
        if len(col) >= 30:
            axis_source += list(col.keys())
    if not axis_source:                              # 全滅時のみ、あるものを全部使う
        for col in table.values():
            axis_source += list(col.keys())
    if not axis_source:
        print("データがまったくありません。既存ファイルを保持して終了します。")
        return 1

    dates = build_date_axis(axis_source)
    table = add_derived(table, dates)

    series, asof = {}, {}
    for key in META:
        col = table.get(key, {})
        series[key] = [latest_on_or_before(col, d) for d in dates] if col else [None] * len(dates)
        asof[key] = max(col.keys()) if col else None

    filled = [k for k in META if asof.get(k)]
    empty = [META[k]["label"] for k in META if not asof.get(k)]
    if empty:
        NOTES.append("未取得: " + " / ".join(empty))

    stamp = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")

    # 系列ごとの「実際に値がある点の数」と「直近5点」。
    # 前方補完のせいで画面上は値があるように見えても、
    # 中身が1点しかない、といった状態をここで見抜けるようにする。
    health = {}
    for key in META:
        col = table.get(key, {})
        recent = [[d, col[d]] for d in sorted(col)[-5:]]
        health[key] = {"points": len(col), "asof": asof[key], "recent": recent}

    payload = {
        # 先頭に診断を置く。ファイルが大きくなっても、頭を読めば状態がわかるようにするため。
        "updated": stamp,
        "years": YEARS,
        "asof": asof,
        "notes": NOTES,
        "meta": META,
        "dates": dates,
        "series": series,
    }
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    # 診断だけの小さなファイル。数値の本体を含まないので軽く、まるごと読める。
    status = {
        "updated": stamp,
        "dates_count": len(dates),
        "dates_first": dates[0],
        "dates_last": dates[-1],
        "elapsed_sec": round(time.time() - t0),
        "size_kb": round(os.path.getsize(OUT_PATH) / 1024.0),
        "missing": empty,
        "notes": NOTES,
        "health": health,
        "diag": DIAG,
    }
    with open(STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=1)

    print("\n" + "=" * 56)
    print("日付 %d 点 (%s 〜 %s)" % (len(dates), dates[0], dates[-1]))
    print("系列 %d 本中 %d 本にデータあり / %.0f KB / %.0f 秒"
          % (len(META), len(filled), os.path.getsize(OUT_PATH) / 1024.0, time.time() - t0))
    thin = [META[k]["label"] for k in META if 0 < health[k]["points"] < 12]
    if thin:
        print("点が少なく判定を出せない系列: " + " / ".join(thin))
    if empty:
        print("取れなかったもの: " + " / ".join(empty))
    print("=" * 56)
    return 0


if __name__ == "__main__":
    sys.exit(main())
